from discord.ext import commands #지울껀가요? 정말요? 지워봐요 ㅋ
import discord #지울 수 있으면 지워봐라
from time import time #쿨타임
from random import * #난수
import openpyxl #엑셀파일
import pandas as pd #엑셀파일 (강화 순위)

bot = commands.Bot(command_prefix='!')

@bot.event
async def on_ready():
    print('Logged in as')
    print(bot.user.name)
    print(bot.user.id)
    print('------')
    global pasttime
    pasttime = dict()

@bot.command(name="강화")
async def 강화(ctx, *args):
    if args == ():
        await ctx.send("강화 순위 - 강화 순위를 확인합니다\n강화 [이름] - [이름]의 레벨을 올립니다 (1~9레벨이 증가하며 파괴확률이 존재합니다)")
    elif args == ('순위',):
        levrank = []
        userrank = []

        ds = pd.read_excel('D:\강화\data.xlsx', sheet_name='Sheet1')

        df = pd.DataFrame(ds)

        df = df.sort_values(by='Level', ascending=False)

        df.to_excel("data.xlsx", index=False)

        wb=None
        ws=None

        wb = openpyxl.load_workbook('data.xlsx')
        ws = wb.active

        for i in range (2, len(ws['B'])+2):
            if i != len(ws['B'])+1:
                levrank.append(int(ws['B'+str(i)].value))

        for v in range (2, len(ws['A'])+2):
            if v != len(ws['A'])+1:
                userrank.append(str(ws['A'+str(v)].value))

        rankst = userrank[0] + ", " + str(levrank[0]) + ".lev"
        ranknd = userrank[1] + ", " + str(levrank[1]) + ".lev"
        rankth = userrank[2] + ", " + str(levrank[2]) + ".lev"
        rankfo = userrank[3] + ", " + str(levrank[3]) + ".lev"
        rankfi = userrank[4] + ", " + str(levrank[4]) + ".lev"
        ranksi = userrank[5] + ", " + str(levrank[5]) + ".lev"
        rankse = userrank[6] + ", " + str(levrank[6]) + ".lev"
        rankei = userrank[7] + ", " + str(levrank[7]) + ".lev"
        rankni = userrank[8] + ", " + str(levrank[8]) + ".lev"
        rankte = userrank[9] + ", " + str(levrank[9]) + ".lev"

        await ctx.send("1위: " + str(rankst) + "\n2위: " + str(ranknd) + "\n3위: " + str(rankth) + "\n4위: " + str(rankfo) + "\n5위: " + str(rankfi) + "\n6위: " + str(ranksi) + "\n7위: " + str(rankse) + "\n8위: " + str(rankei) + "\n9위: " + str(rankni) + "\n10위: " + str(rankte))
    else:
        if len(str(args)) < 20:
            if ctx.author.id not in pasttime:
                pasttime[ctx.author.id] = time()
            else:
                past = pasttime[ctx.author.id]
                if time() - past < 60:
                    em = discord.Embed(title=f"아직 강화를 진행할 수 없습니다! " + str(round((float(60) - (float(time()) - float(past))), 2)) + " 초 이후에 다시 시도해주세요.", color=15158332)
                    await ctx.send(embed=em)
                    return
            pasttime[ctx.author.id] = time()

            wb = openpyxl.load_workbook('data.xlsx')
            ws = wb.active

            argt = str(args)[2:]
            argtr = argt.find("'")
            argf = argt[:argtr]

            for i in range (2, len(ws['A'])+2):
                if i != len(ws['A'])+1:
                    if str(ws['A'+str(i)].value) == str(argf):
                        lev = int(ws['B'+str(i)].value)
                        break
                else:
                    lev = 0
                    break

            randf = uniform(95.0-(lev*0.613), 99.9)
            randtw = randrange(0, 100)

            if randtw < randf:
                an = True
            else:
                an = False

            if an == True:
                rand = randrange(1,10)
                if i != len(ws['A'])+1:
                    await ctx.send(str(argf) + " 를 " + str(round(randf, 2)) + "%의 확률로 강화에 성공하였습니다\n" + str(argf) + " 의 레벨 " + str(lev) + "->" + str(int(lev) + int(rand)))
                    lev = lev + rand
                    ws['B'+str(i)] = lev
                    wb.save('data.xlsx')
                else:
                    ws['A'+str(i)] = str(argf)
                    ws['B'+str(i)] = rand
                    wb.save('data.xlsx')
                    await ctx.send(str(argf) + " 를 " + str(round(randf, 2)) + "%의 확률로 강화에 성공하였습니다\n" + str(argf) + " 의 레벨 " + str(rand))
            else:
                rand = randrange(1, 15)
                if i != len(ws['A'])+1:
                    randt = uniform(95.0-(lev*0.05), 99.9)
                    if randt > randtw:
                        await ctx.send(str(argf) + " 가 " + str(100-round(randt, 2)) + "%의 확률로 파괴되었습니다\n" + str(argf) + " 의 레벨 " + str(lev) + "->0")
                        ws.delete_rows(i)
                        wb.save('data.xlsx')
                    else:
                        await ctx.send(str(argf) + " 가 " + str(100-round(randf, 2)) + "%의 확률로 강화에 실패하였습니다\n" + str(argf) + " 의 레벨 " + str(lev) + "->" + str(int(lev) - int(rand)))
                        lev = lev - rand
                        ws['B'+str(i)] = lev
                        wb.save('data.xlsx')
                else:
                    await ctx.send(str(argf) + " 가 " + str(100-round(randf, 2)) + "%의 확률로 강화에 실패하였습니다\n" + str(argf) + "가 존재하지 않아 0레벨로 설정되었습니다")
        else:
            await ctx.send("이름이 너무 깁니다")

bot.run('TOKEN')